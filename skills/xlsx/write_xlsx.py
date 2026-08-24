#!/usr/bin/env python3
"""Create an Excel .xlsx from CSV or JSON.  No system deps (openpyxl).

Usage:  write_xlsx.py <input.csv>  <output.xlsx>      # single sheet from CSV
        write_xlsx.py <input.json> <output.xlsx>      # one or more sheets
        write_xlsx.py - <output.xlsx> < data.json     # read from stdin
Venv:   pip install openpyxl

JSON forms accepted:
  {"Sheet1": [["h1","h2"], [1,2], [3,4]], "Sheet2": [...]}      # name -> rows
  [{"name": "Sheet1", "rows": [[...], ...]}, ...]               # list form
The first row of each sheet is treated as a bold header.
"""
import csv
import json
import sys


def _sheets(src):
    raw = sys.stdin.read() if src == "-" else open(src, encoding="utf-8").read()
    if src.endswith(".json") or (src == "-" and raw.lstrip()[:1] in "[{"):
        data = json.loads(raw)
        if isinstance(data, dict):
            return [(k, v) for k, v in data.items()]
        return [(d.get("name", f"Sheet{i+1}"), d.get("rows", []))
                for i, d in enumerate(data)]
    rows = list(csv.reader(raw.splitlines()))
    return [("Sheet1", rows)]


def main(src, out):
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    wb = Workbook(); wb.remove(wb.active)
    for name, rows in _sheets(src):
        ws = wb.create_sheet(title=str(name)[:31] or "Sheet")
        widths = {}
        for r, row in enumerate(rows, 1):
            for c, val in enumerate(row, 1):
                cell = ws.cell(row=r, column=c, value=val)
                if r == 1:
                    cell.font = Font(bold=True)
                widths[c] = max(widths.get(c, 0), len(str(val)) if val is not None else 0)
        for c, w in widths.items():
            ws.column_dimensions[get_column_letter(c)].width = min(max(w + 2, 8), 60)
        if rows:
            ws.freeze_panes = "A2"     # keep header visible
    if not wb.sheetnames:
        wb.create_sheet("Sheet1")
    wb.save(out)
    print(f"wrote {out} ({len(wb.sheetnames)} sheet(s))")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        sys.exit("usage: write_xlsx.py <input.csv|input.json|-> <output.xlsx>")
    main(sys.argv[1], sys.argv[2])
